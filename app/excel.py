import xlsxwriter
import openpyxl

from .config import file_excel
from .cell_formats import *

def adjust_column_width(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    column_max_length = [0] * sheet.max_column

    ya_ustal = 0
    for row in sheet.iter_rows(values_only=True):
        if (ya_ustal != 0):
            for idx, cell_value in enumerate(row):
                width = 0
                if (cell_value is not None):
                    strs = str(cell_value).split('\n')
                    for i in strs:
                        if (len(i) > width):
                            width = len(i)

                    cell_length = width
                    if cell_length > column_max_length[idx]:
                        column_max_length[idx] = cell_length
        ya_ustal += 1
    for col_idx, max_length in enumerate(column_max_length[1:], start=2):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 10 if max_length < 10 else max_length

    workbook.save(filename)


def save_full_to_excel(wallets, chains, coins, balances):
    workbook = xlsxwriter.Workbook(file_excel)
    worksheet = workbook.add_worksheet("Coins")

    header_format = workbook.add_format(header_format_dict)
    wallets_column_format = workbook.add_format(wallets_column_format_dict)
    total_cell_format = workbook.add_format(total_cell_format_dict)
    common_ceil_format = workbook.add_format(common_ceil_format_dict)
    usd_ceil_format = workbook.add_format(usd_ceil_format_dict)
    donate_cell_format = workbook.add_format(donate_cell_format_dict)

    headers = ['Wallet', *[chain.upper() for chain in chains], 'CHAINS', 'TOTAL']

    for row_id, wallet in enumerate(wallets):
        worksheet.write(row_id + 1, 0, wallet, wallets_column_format)
    worksheet.write(len(wallets) + 1, 0, 'TOTAL IN USD', total_cell_format)


    for col_id, chain in enumerate(headers):
        worksheet.write(0, col_id, chain, header_format)


    for col_id, chain in enumerate(chains):
        total_in_chain = 0.0
        for row_id, wallet in enumerate(wallets):
            cell = ''
            for coin in coins[chain][wallet]:
                coin_in_usd = '?' if (coin["price"] is None) else round(coin["amount"] * coin["price"], 2)
                cell += f'{coin["ticker"]} - {round(coin["amount"], 4)} (${coin_in_usd})\n'
                total_in_chain += coin_in_usd if (type(coin_in_usd) is float) else 0
            if (cell == ''):
                cell = '--'
            cell = cell[:-1]
            worksheet.write(row_id + 1, col_id + 1, cell, common_ceil_format)
        worksheet.write(len(wallets) + 1, col_id + 1, f'${round(total_in_chain, 2)}', usd_ceil_format)


    total_usd = 0.0
    total_all_chains = 0.0
    for row_id, wallet in enumerate(wallets):
        total_in_wallet = 0.0
        for chain in chains:
            for coin in coins[chain][wallet]:
                coin_in_usd = 0 if (coin["price"] is None) else round(coin["amount"] * coin["price"], 2)
                total_in_wallet += coin_in_usd
        total_usd += total_in_wallet
        total_all_chains += balances[wallet]
        worksheet.write(row_id + 1, len(headers) - 2, f'${round(total_in_wallet, 2)}', usd_ceil_format)
        worksheet.write(row_id + 1, len(headers) - 1, f'${round(balances[wallet], 2)}', usd_ceil_format)
    worksheet.write(len(wallets) + 1, len(headers) - 2, f'${round(total_usd, 2)}', usd_ceil_format)
    worksheet.write(len(wallets) + 1, len(headers) - 1, f'${round(total_all_chains, 2)}', usd_ceil_format)


    worksheet.write(len(wallets) + 3, 0, 'Donate:', donate_cell_format)
    worksheet.write(len(wallets) + 4, 0, '0x2e69Da32b0F7e75549F920CD2aCB0532Cc2aF0E7', donate_cell_format)


    worksheet.set_row(0, 35)
    worksheet.set_column(0, 0, 52)

    workbook.close()

    adjust_column_width(file_excel)


def save_selected_to_excel(wallets, chains, coins, balances, ticker):
    workbook = xlsxwriter.Workbook(file_excel)
    worksheet = workbook.add_worksheet("Coins")

    header_format = workbook.add_format(header_format_dict)
    wallets_column_format = workbook.add_format(wallets_column_format_dict)
    total_cell_format = workbook.add_format(total_cell_format_dict)
    common_ceil_format = workbook.add_format(common_ceil_format_dict)
    usd_ceil_format = workbook.add_format(usd_ceil_format_dict)
    donate_cell_format = workbook.add_format(donate_cell_format_dict)

    headers = ['Wallet', *[chain.upper() for chain in chains], 'CHAINS', 'TOTAL']

    for row_id, wallet in enumerate(wallets):
        worksheet.write(row_id + 1, 0, wallet, wallets_column_format)
    worksheet.write(len(wallets) + 1, 0, 'TOTAL IN USD', total_cell_format)


    for col_id, chain in enumerate(headers):
        if (col_id == 0):
            worksheet.write(0, 0, chain, header_format)
        elif (col_id in [len(headers) - 1, len(headers) - 2]):
            worksheet.write(0, col_id + (len(headers) - 3) * 2, chain, header_format)
        else:
            worksheet.merge_range(0, col_id - 2 + 2 * col_id, 0, col_id + 2 * col_id, chain, header_format)


    for col_id, chain in enumerate(chains):
        total_in_chain = 0.0
        total_amount = 0.0
        for row_id, wallet in enumerate(wallets):
            if (ticker in [coin['ticker'] for coin in coins[chain][wallet]]):
                for coin in coins[chain][wallet]:
                    if (coin['ticker'] == ticker):
                        coin_in_usd = '?' if (coin["price"] is None) else round(coin["amount"] * coin["price"], 2)
                        total_in_chain += coin_in_usd if (type(coin_in_usd) is float) else 0
                        total_amount += coin['amount']
                        worksheet.write(row_id + 1, col_id - 1 + (col_id + 1) * 2, ticker, common_ceil_format)
                        worksheet.write(row_id + 1, col_id + (col_id + 1) * 2, round(coin['amount'], 4), common_ceil_format)
                        worksheet.write(row_id + 1, col_id + 1 + (col_id + 1) * 2, f'${coin_in_usd}', common_ceil_format)
            else:

                worksheet.write(row_id + 1, col_id - 1 + (col_id + 1) * 2, ticker, common_ceil_format)
                worksheet.write(row_id + 1, col_id + (col_id + 1) * 2, 0, common_ceil_format)
                worksheet.write(row_id + 1, col_id + 1 + (col_id + 1) * 2, '$0', common_ceil_format)
        worksheet.write(len(wallets) + 1, col_id - 1 + (col_id + 1) * 2, ticker, usd_ceil_format)
        worksheet.write(len(wallets) + 1, col_id + (col_id + 1) * 2, round(total_amount, 4), usd_ceil_format)
        worksheet.write(len(wallets) + 1, col_id + 1 + (col_id + 1) * 2, f'${round(total_in_chain, 2)}', usd_ceil_format)


    total_usd = 0.0
    total_all_chains = 0.0
    for row_id, wallet in enumerate(wallets):
        total_in_wallet = 0.0
        for chain in chains:
            for coin in coins[chain][wallet]:
                if (coin['ticker'] == ticker):
                    coin_in_usd = 0 if (coin["price"] is None) else round(coin["amount"] * coin["price"], 2)
                    total_in_wallet += coin_in_usd
        total_usd += total_in_wallet
        total_all_chains += balances[wallet]
        worksheet.write(row_id + 1, len(headers) - 2 + (len(headers) - 3) * 2, f'${round(total_in_wallet, 2)}', usd_ceil_format)
        worksheet.write(row_id + 1, len(headers) - 1 + (len(headers) - 3) * 2, f'${round(balances[wallet], 2)}', usd_ceil_format)
    worksheet.write(len(wallets) + 1, len(headers) - 2 + (len(headers) - 3) * 2, f'${round(total_usd, 2)}', usd_ceil_format)
    worksheet.write(len(wallets) + 1, len(headers) - 1 + (len(headers) - 3) * 2, f'${round(total_all_chains, 2)}', usd_ceil_format)


    worksheet.write(len(wallets) + 3, 0, 'Donate:', donate_cell_format)
    worksheet.write(len(wallets) + 4, 0, '0x2e69Da32b0F7e75549F920CD2aCB0532Cc2aF0E7', donate_cell_format)


    worksheet.set_row(0, 35)
    worksheet.set_column(0, 0, 52)

    workbook.close()

    adjust_column_width(file_excel)
